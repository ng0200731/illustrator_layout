"""
Tool to read and analyze the actual Excel file to understand the data exchange issue.
"""
import sys
import os

def read_excel_file(file_path):
    """Read and analyze the Excel file."""
    try:
        import openpyxl

        print(f"=== Reading Excel File: {file_path} ===")

        wb = openpyxl.load_workbook(file_path, read_only=True)
        print(f"Worksheets: {wb.sheetnames}")

        # Read the main data sheet
        ws = wb['Data'] if 'Data' in wb.sheetnames else wb.active
        print(f"Active sheet: {ws.title}")

        # Read headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))

        print(f"Headers: {headers}")

        # Read data rows
        data_rows = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if any(cell is not None for cell in row):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                data_rows.append({
                    'row_num': row_num,
                    'data': row_data
                })

        print(f"Found {len(data_rows)} data rows")

        # Show the data
        for i, row in enumerate(data_rows):
            print(f"\nRow {row['row_num']}:")
            for j, header in enumerate(headers):
                if j < len(row['data']):
                    value = row['data'][j]
                    if len(value) > 60:
                        print(f"  {header}: {value[:60]}...")
                    else:
                        print(f"  {header}: {value}")

        # Check for metadata sheet
        if '_metadata' in wb.sheetnames:
            print(f"\n=== Metadata Sheet Found ===")
            meta_ws = wb['_metadata']
            for row in meta_ws.iter_rows(values_only=True):
                if row[0] is not None:
                    print(f"  {row}")
        else:
            print(f"\n=== No Metadata Sheet ===")

        wb.close()
        return headers, data_rows

    except ImportError:
        print("openpyxl not available. Install with: pip install openpyxl")
        return None, None
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None, None

def analyze_variable_mapping_issue(headers, data_rows):
    """Analyze why the variable mapping isn't working."""
    if not headers or not data_rows:
        return

    print(f"\n=== Variable Mapping Analysis ===")

    # Expected mapping based on your description
    expected_mapping = {
        0: {"header": "barcode_digits", "var_id": 15, "expected": "8447692183702"},
        1: {"header": "QR", "var_id": 26, "expected": "https://qr.mango.com/...6074"},
        2: {"header": "barcode_graphic", "var_id": 27, "expected": "8447542484929"}
    }

    print("Expected mapping:")
    for col_idx, mapping in expected_mapping.items():
        print(f"  Column {col_idx} ({mapping['header']}) → Variable #{mapping['var_id']}")
        print(f"    Expected value: {mapping['expected']}")

    print(f"\nActual data in Excel:")
    if data_rows:
        row = data_rows[0]  # First data row
        for i, header in enumerate(headers):
            if i < len(row['data']):
                value = row['data'][i]
                expected_var = None
                for col_idx, mapping in expected_mapping.items():
                    if col_idx == i:
                        expected_var = mapping['var_id']
                        break

                if len(value) > 60:
                    print(f"  Column {i} ({header}) → Variable #{expected_var}: {value[:60]}...")
                else:
                    print(f"  Column {i} ({header}) → Variable #{expected_var}: {value}")

def create_debug_order_payload(headers, data_rows):
    """Create the exact order payload that should be generated."""
    if not headers or not data_rows:
        return

    print(f"\n=== Debug Order Payload ===")

    # Map column indices to variable IDs
    column_to_variable = {
        0: 15,  # barcode_digits → Variable #15
        1: 26,  # QR → Variable #26
        2: 27   # barcode_graphic → Variable #27
    }

    for row in data_rows:
        print(f"\nRow {row['row_num']}:")
        variable_values = {}
        qty = 1

        for col_idx, value in enumerate(row['data']):
            if col_idx in column_to_variable:
                var_id = column_to_variable[col_idx]
                variable_values[str(var_id)] = value
                print(f"  Variable #{var_id}: {value[:50]}..." if len(value) > 50 else f"  Variable #{var_id}: {value}")
            elif col_idx < len(headers) and headers[col_idx].lower() == 'qty':
                try:
                    qty = int(float(value))
                    print(f"  Quantity: {qty}")
                except:
                    print(f"  Quantity: {value} (invalid)")

        # This is what the order system should create
        order_row = {
            "variableValues": variable_values,
            "quantity": qty
        }

        print(f"  Order payload: {order_row}")

def main():
    """Main function to analyze the Excel file."""
    file_path = r"C:\Users\ng\Downloads\test.xlsx"

    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    headers, data_rows = read_excel_file(file_path)
    analyze_variable_mapping_issue(headers, data_rows)
    create_debug_order_payload(headers, data_rows)

    print(f"\n=== DIAGNOSIS ===")
    print("If the Excel data looks correct but variables still show wrong values:")
    print("1. The Excel processing code isn't using the correct column mapping")
    print("2. The metadata sheet is missing or has wrong variable indices")
    print("3. The layout variables don't match the expected IDs (15,26,27)")
    print("4. There's caching preventing the updated code from running")

if __name__ == "__main__":
    main()