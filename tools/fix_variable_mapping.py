"""
Tool to fix variable mapping issues where row-specific data isn't preserved.

The core issue: Variable indices don't match between Excel columns and layout variables.
This causes all rows to use the same variable values instead of row-specific data.
"""
import json
import os
import sys

def analyze_variable_index_mapping():
    """Analyze how variable indices should map to Excel columns."""
    print("=== Variable Index Mapping Analysis ===")

    # Your layout uses these variables:
    layout_variables = [
        {"idx": 15, "content": "barcode_digits", "variableName": "barcode_digits"},
        {"idx": 26, "content": "qr_data", "variableName": "qr_data"},
        {"idx": 27, "content": "barcode_display", "variableName": "barcode_display"}
    ]

    print("Layout variables:")
    for var in layout_variables:
        print(f"  Variable #{var['idx']}: {var['content']}")

    # Excel columns (0-based indexing)
    excel_columns = [
        {"col": 0, "header": "barcode_digits"},
        {"col": 1, "header": "qr_data"},
        {"col": 2, "header": "barcode_display"},
        {"col": 3, "header": "qty"}
    ]

    print("\nExcel columns:")
    for col in excel_columns:
        print(f"  Column {col['col']}: {col['header']}")

    print("\n=== THE PROBLEM ===")
    print("Excel processing maps columns 0,1,2 to variable indices 0,1,2")
    print("But your layout expects variable indices 15,26,27")
    print("This mismatch causes wrong variable assignment!")

    return layout_variables, excel_columns

def create_metadata_fix():
    """Create Excel with proper metadata to fix variable mapping."""
    print("\n=== Creating Fixed Excel with Metadata ===")

    try:
        import openpyxl

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Headers
        ws.cell(row=1, column=1, value='barcode_digits')
        ws.cell(row=1, column=2, value='qr_data')
        ws.cell(row=1, column=3, value='barcode_display')
        ws.cell(row=1, column=4, value='qty')

        # Your actual data
        data_rows = [
            ["8447692183473", "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116073", "8447692183473", 11],
            ["8447692183702", "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116074", "8447692183702", 22]
        ]

        for i, row_data in enumerate(data_rows, start=2):
            for j, value in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=value)

        # Create metadata sheet with correct variable mapping
        meta = wb.create_sheet("_metadata")
        meta.sheet_state = 'hidden'
        meta.cell(row=1, column=1, value="col_position")
        meta.cell(row=1, column=2, value="variable_index")
        meta.cell(row=1, column=3, value="default_content")

        # Map Excel columns to correct variable indices
        variable_mapping = [
            {"col_pos": 1, "var_idx": 15, "content": "barcode_digits"},
            {"col_pos": 2, "var_idx": 26, "content": "qr_data"},
            {"col_pos": 3, "var_idx": 27, "content": "barcode_display"}
        ]

        for i, mapping in enumerate(variable_mapping, start=2):
            meta.cell(row=i, column=1, value=mapping["col_pos"])
            meta.cell(row=i, column=2, value=mapping["var_idx"])
            meta.cell(row=i, column=3, value=mapping["content"])

        output_file = ".tmp/mango_fixed_variables.xlsx"
        wb.save(output_file)

        print(f"Created fixed Excel: {output_file}")
        print("This file has correct variable index mapping in metadata")

        return True

    except ImportError:
        print("openpyxl not available, creating manual CSV with instructions")
        return create_manual_csv_fix()
    except Exception as e:
        print(f"Error creating Excel: {e}")
        return False

def create_manual_csv_fix():
    """Create CSV with instructions for manual variable mapping."""
    print("\n=== Manual CSV Fix ===")

    csv_content = """# IMPORTANT: Variable Mapping Instructions
# Column 1 (barcode_digits) -> Variable #15
# Column 2 (qr_data) -> Variable #26
# Column 3 (barcode_display) -> Variable #27
#
# Make sure the system maps these correctly!
barcode_digits,qr_data,barcode_display,qty
8447692183473,https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116073,8447692183473,11
8447692183702,https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116074,8447692183702,22
"""

    try:
        with open(".tmp/mango_manual_fix.csv", 'w', encoding='utf-8') as f:
            f.write(csv_content)
        print("Created manual CSV: .tmp/mango_manual_fix.csv")
        return True
    except Exception as e:
        print(f"Error creating CSV: {e}")
        return False

def debug_order_processing():
    """Debug how the order system should process the data."""
    print("\n=== Order Processing Debug ===")

    # Simulate correct processing
    rows = [
        {
            "variableValues": {
                "15": "8447692183473",  # barcode_digits for row 1
                "26": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116073",  # qr_data for row 1
                "27": "8447692183473"   # barcode_display for row 1
            },
            "quantity": 11
        },
        {
            "variableValues": {
                "15": "8447692183702",  # barcode_digits for row 2
                "26": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116074",  # qr_data for row 2
                "27": "8447692183702"   # barcode_display for row 2
            },
            "quantity": 22
        }
    ]

    print("Correct order processing should create:")
    for i, row in enumerate(rows, 1):
        print(f"\nRow {i} (qty: {row['quantity']}):")
        for var_id, value in row['variableValues'].items():
            if len(str(value)) > 60:
                print(f"  Variable #{var_id}: {str(value)[:60]}...")
            else:
                print(f"  Variable #{var_id}: {value}")

    # Save as JSON for testing
    with open(".tmp/correct_order_payload.json", 'w', encoding='utf-8') as f:
        json.dump({"rows": rows}, f, indent=2)

    print(f"\nSaved correct payload: .tmp/correct_order_payload.json")

def check_layout_for_hardcoded_values():
    """Check if the layout has hardcoded values that override variables."""
    print("\n=== Layout Hardcoded Values Check ===")
    print("The issue might be:")
    print("1. Layout has hardcoded barcode '8447692183949' instead of using variable #15")
    print("2. QR code generation uses fixed base URL instead of variable #26")
    print("3. Variable substitution happens after hardcoded values are set")
    print("\nTo fix:")
    print("1. Check layout JSON for hardcoded '8447692183949'")
    print("2. Ensure barcode elements use variable #15, not fixed values")
    print("3. Ensure QR elements use variable #26, not base URL")

def main():
    """Main function to fix variable mapping."""
    print("=== Variable Mapping Fix Tool ===")

    analyze_variable_index_mapping()
    create_metadata_fix()
    debug_order_processing()
    check_layout_for_hardcoded_values()

    print("\n=== SOLUTION STEPS ===")
    print("1. Use the fixed Excel file: .tmp/mango_fixed_variables.xlsx")
    print("2. Check that variable mapping uses indices 15,26,27 (not 0,1,2)")
    print("3. Verify layout doesn't have hardcoded values")
    print("4. Test with the correct order payload JSON")

if __name__ == "__main__":
    main()