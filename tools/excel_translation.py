"""Excel parsing and export for translation tables."""
import openpyxl

def parse_translation_excel(filepath):
    """Parse Excel file to JSON structure.

    Args:
        filepath: Path to Excel file

    Returns:
        dict with 'headers' and 'rows':
        {
            "headers": ["Material", "Spanish (ES)", "French (FR)", ...],
            "rows": [
                ["COTTON", "algodón", "coton", ...],
                ["POLYESTER", "poliéster", "polyester", ...]
            ]
        }
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        # Use first/active sheet only
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)

        # Extract headers from row 1
        header_row = next(rows_iter)
        headers = [str(h).strip() if h is not None else '' for h in header_row]

        # Remove trailing empty headers
        while headers and headers[-1] == '':
            headers.pop()

        if not headers:
            raise ValueError("No headers found in Excel file")

        # Extract data rows (skip empty rows)
        data_rows = []
        for row in rows_iter:
            # Convert row to list and trim to header length
            row_values = list(row[:len(headers)])

            # Check if row is empty
            is_empty = all(v is None or str(v).strip() == '' for v in row_values)
            if is_empty:
                continue

            # Convert all values to strings
            row_data = [str(v).strip() if v is not None else '' for v in row_values]
            data_rows.append(row_data)

        if not data_rows:
            raise ValueError("No data rows found in Excel file")

        return {
            "headers": headers,
            "rows": data_rows
        }
    finally:
        wb.close()

def export_translation_excel(data, output_path):
    """Export JSON structure to Excel file.

    Args:
        data: dict with 'headers' and 'rows'
        output_path: Path to save Excel file

    Returns:
        output_path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"

    # Write headers
    headers = data.get('headers', [])
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    rows = data.get('rows', [])
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    wb.save(output_path)
    return output_path
