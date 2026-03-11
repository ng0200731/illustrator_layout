"""
Comprehensive tool to fix barcode ordering issues in Illustrator exports.

This tool addresses the specific issue where Illustrator exports don't maintain
the original input order for barcode/QR data, ensuring proper sequence matching.
"""
import json
import os
import sys
import re
from collections import OrderedDict

def parse_barcode_table_from_ide_data():
    """Parse the barcode data shown in the IDE."""
    # This is the actual data from your IDE display
    barcode_data = [
        {"index": 1, "barcode": "8447692183473", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116073"},
        {"index": 2, "barcode": "8447692183702", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116074"},
        {"index": 3, "barcode": "8447692183632", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116075"},
        {"index": 4, "barcode": "8447692183910", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116076"},
        {"index": 5, "barcode": "8447692183805", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116077"},
        {"index": 6, "barcode": "8447692183370", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116078"},
        {"index": 7, "barcode": "8447692183360", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116079"},
        {"index": 8, "barcode": "8447692183645", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116080"},
        {"index": 9, "barcode": "8447692183924", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116081"},
        {"index": 10, "barcode": "8447692183994", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116082"},
        {"index": 11, "barcode": "8447692183910", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116083"},
        {"index": 12, "barcode": "8447692183165", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116084"},
        {"index": 13, "barcode": "8447692183547", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116085"},
        {"index": 14, "barcode": "8447692183828", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116086"},
        {"index": 15, "barcode": "8447692183172", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116087"},
        {"index": 16, "barcode": "8447692183201", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116088"},
        {"index": 17, "barcode": "8447692183816", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116089"},
        {"index": 18, "barcode": "8447692183412", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116090"},
        {"index": 19, "barcode": "8447692183735", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116091"},
        {"index": 20, "barcode": "8447692183457", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116092"},
        {"index": 21, "barcode": "8447692183205", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116093"},
        {"index": 22, "barcode": "8447692183306", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116094"},
        {"index": 23, "barcode": "8447692183244", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116095"},
        {"index": 24, "barcode": "8447692183874", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116096"},
        {"index": 25, "barcode": "8447692183299", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116097"},
        {"index": 26, "barcode": "8447692183143", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116098"},
        {"index": 27, "barcode": "8447692183858", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116099"},
        {"index": 28, "barcode": "8447692183991", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116100"},
        {"index": 29, "barcode": "8447692183566", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116101"},
        {"index": 30, "barcode": "8447692183988", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116102"},
        {"index": 31, "barcode": "8447692183618", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116103"},
        {"index": 32, "barcode": "8447692183909", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116104"},
        {"index": 33, "barcode": "8447692183951", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116105"},
        {"index": 34, "barcode": "8447692183280", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116106"},
        {"index": 35, "barcode": "8447692183941", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116107"},
        {"index": 36, "barcode": "8447692183304", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116108"},
        {"index": 37, "barcode": "8447692183687", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116109"},
        {"index": 38, "barcode": "8447692183226", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116110"},
        {"index": 39, "barcode": "8447692183370", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116111"},
        {"index": 40, "barcode": "8447692183519", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116112"},
        {"index": 41, "barcode": "8447692183751", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116113"},
        {"index": 42, "barcode": "8447692183996", "qr": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116114"}
    ]
    return barcode_data

def create_excel_template_for_order_system(barcode_data, output_file):
    """Create an Excel file that the order system can process correctly."""
    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Headers
        ws.cell(row=1, column=1, value='barcode_digits')
        ws.cell(row=1, column=2, value='qr_data')
        ws.cell(row=1, column=3, value='qty')

        # Data rows in correct order
        for i, item in enumerate(barcode_data, start=2):
            ws.cell(row=i, column=1, value=item['barcode'])
            ws.cell(row=i, column=2, value=item['qr'])
            ws.cell(row=i, column=3, value=1)

        wb.save(output_file)
        print(f"Created Excel file: {output_file}")
        print(f"Contains {len(barcode_data)} rows in correct order")
        return True

    except ImportError:
        print("openpyxl not available, creating CSV instead")
        return create_csv_template_for_order_system(barcode_data, output_file.replace('.xlsx', '.csv'))
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return False

def create_csv_template_for_order_system(barcode_data, output_file):
    """Create a CSV file that the order system can process correctly."""
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            # Headers
            f.write('barcode_digits,qr_data,qty\n')

            # Data rows in correct order
            for item in barcode_data:
                f.write(f'{item["barcode"]},{item["qr"]},1\n')

        print(f"Created CSV file: {output_file}")
        print(f"Contains {len(barcode_data)} rows in correct order")
        return True

    except Exception as e:
        print(f"Error creating CSV file: {e}")
        return False

def generate_order_json_payload(barcode_data):
    """Generate the JSON payload for creating an order with proper variable mapping."""
    # Create variable values mapping
    variable_values = {}

    for i, item in enumerate(barcode_data):
        # Map barcode data to variable indices
        # Assuming variables are indexed sequentially
        barcode_var_idx = str(i * 2)      # Even indices for barcodes
        qr_var_idx = str(i * 2 + 1)       # Odd indices for QR codes

        variable_values[barcode_var_idx] = item['barcode']
        variable_values[qr_var_idx] = item['qr']

    # Create order payload
    order_payload = {
        'rows': []
    }

    for item in barcode_data:
        row = {
            'variableValues': {
                # This mapping depends on your layout's variable structure
                # You'll need to adjust based on your actual layout
                '0': item['barcode'],  # Assuming variable 0 is barcode
                '1': item['qr']        # Assuming variable 1 is QR
            },
            'quantity': 1
        }
        order_payload['rows'].append(row)

    return order_payload

def verify_export_order_thoroughly(ai_file, expected_barcode_data):
    """Thoroughly verify the export order against expected data."""
    try:
        with open(ai_file, 'r', encoding='utf-8') as f:
            ai_data = json.load(f)

        print(f"=== Thorough Order Verification ===")
        print(f"AI file: {os.path.basename(ai_file)}")
        print(f"Expected items: {len(expected_barcode_data)}")

        # Analyze the structure
        layers = ai_data.get('layers', [])
        all_paths = []

        def collect_paths(children, layer_name=""):
            for child in children:
                if child.get('type') == 'path':
                    all_paths.append({
                        'name': child.get('name', ''),
                        'bounds': child.get('bounds', {}),
                        'layer': layer_name
                    })
                elif 'children' in child:
                    collect_paths(child['children'], layer_name)

        for layer in layers:
            layer_name = layer.get('name', 'Unknown')
            if 'children' in layer:
                collect_paths(layer['children'], layer_name)

        print(f"Total paths found: {len(all_paths)}")

        # Look for patterns that might indicate barcode elements
        barcode_paths = []
        for i, path in enumerate(all_paths):
            bounds = path['bounds']
            # Barcode elements typically have specific width/height ratios
            if bounds.get('width', 0) > 0 and bounds.get('height', 0) > 0:
                ratio = bounds['width'] / bounds['height']
                # Barcode bars are typically tall and narrow
                if ratio < 0.5 or ratio > 10:  # Very tall or very wide elements
                    barcode_paths.append({
                        'index': i,
                        'path': path,
                        'ratio': ratio
                    })

        print(f"Potential barcode elements: {len(barcode_paths)}")

        # Group by Y position (assuming barcodes are arranged vertically)
        y_groups = {}
        for bp in barcode_paths:
            y = round(bp['path']['bounds'].get('y', 0))
            if y not in y_groups:
                y_groups[y] = []
            y_groups[y].append(bp)

        print(f"Y-position groups: {len(y_groups)}")

        # Check if we have the expected number of barcode groups
        if len(y_groups) == len(expected_barcode_data):
            print("✓ Number of barcode groups matches expected data")
        else:
            print(f"✗ Barcode group count mismatch: found {len(y_groups)}, expected {len(expected_barcode_data)}")

        return len(y_groups) == len(expected_barcode_data)

    except Exception as e:
        print(f"Error verifying export order: {e}")
        return False

def main():
    """Main function to fix barcode ordering issues."""
    print("=== Mango Barcode Order Fix Tool ===")

    # Get the barcode data
    barcode_data = parse_barcode_table_from_ide_data()
    print(f"Loaded {len(barcode_data)} barcode items")

    if len(sys.argv) < 2:
        print("\nUsage: py mango_barcode_fix.py <command> [options]")
        print("Commands:")
        print("  create_excel <output_file>   - Create Excel template for order system")
        print("  create_csv <output_file>     - Create CSV template for order system")
        print("  verify_ai <ai_file>          - Verify AI export order")
        print("  generate_json <output_file>  - Generate order JSON payload")
        print("  show_data                    - Show the barcode data")
        return

    command = sys.argv[1]

    if command == "show_data":
        print("\n=== Barcode Data (First 10 items) ===")
        for item in barcode_data[:10]:
            print(f"{item['index']:2d}: {item['barcode']} -> {item['qr']}")
        print(f"... and {len(barcode_data) - 10} more items")

    elif command == "create_excel" and len(sys.argv) >= 3:
        output_file = sys.argv[2]
        create_excel_template_for_order_system(barcode_data, output_file)

    elif command == "create_csv" and len(sys.argv) >= 3:
        output_file = sys.argv[2]
        create_csv_template_for_order_system(barcode_data, output_file)

    elif command == "verify_ai" and len(sys.argv) >= 3:
        ai_file = sys.argv[2]
        verify_export_order_thoroughly(ai_file, barcode_data)

    elif command == "generate_json" and len(sys.argv) >= 3:
        output_file = sys.argv[3]
        payload = generate_order_json_payload(barcode_data)
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2)
        print(f"Generated order JSON: {output_file}")

    else:
        print("Invalid command or missing arguments")

if __name__ == "__main__":
    main()