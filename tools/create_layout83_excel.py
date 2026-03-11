"""
Create the CORRECT Excel file that matches the layout indices.
"""
import openpyxl

def create_correct_excel_for_layout():
    """Create Excel with metadata that matches Layout 83 indices."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Headers
    ws.cell(row=1, column=1, value='barcode_digits')
    ws.cell(row=1, column=2, value='QR')
    ws.cell(row=1, column=3, value='barcode_graphic')
    ws.cell(row=1, column=4, value='qty')

    # Sample data with DIFFERENT values for barcode_graphic
    sample_data = [
        ["8447692183473", "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116073", "8447542484929", 1],
        ["8447692183702", "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116074", "8447542484930", 1],
        ["8447692183632", "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116075", "8447542484931", 1]
    ]

    for i, row_data in enumerate(sample_data, start=2):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    # Create metadata sheet with CORRECT indices for Layout 83
    meta = wb.create_sheet("_metadata")
    meta.sheet_state = 'hidden'
    meta.cell(row=1, column=1, value="col_position")
    meta.cell(row=1, column=2, value="variable_index")
    meta.cell(row=1, column=3, value="default_content")

    # CORRECT mapping for Layout 83 (indices 14,25,26 not 15,26,27)
    correct_mappings = [
        (1, 14, "barcode_digits"),    # Column 1 → Index 14 (barcode_digits)
        (2, 25, "QR"),               # Column 2 → Index 25 (QR)
        (3, 26, "barcode_graphic")   # Column 3 → Index 26 (barcode_graphic)
    ]

    for i, (col_pos, var_idx, content) in enumerate(correct_mappings, start=2):
        meta.cell(row=i, column=1, value=col_pos)
        meta.cell(row=i, column=2, value=var_idx)
        meta.cell(row=i, column=3, value=content)

    output_file = "C:/Users/ng/Downloads/test_layout83_fixed.xlsx"
    wb.save(output_file)

    print(f"Created Excel for Layout 83: {output_file}")
    print("Metadata mapping:")
    print("  Column 1 (barcode_digits) → Variable Index #14")
    print("  Column 2 (QR) → Variable Index #25")
    print("  Column 3 (barcode_graphic) → Variable Index #26")
    print()
    print("Sample data:")
    print("  Row 1: barcode_digits=8447692183473, barcode_graphic=8447542484929 (DIFFERENT!)")
    print("  Row 2: barcode_digits=8447692183702, barcode_graphic=8447542484930 (DIFFERENT!)")
    print("  Row 3: barcode_digits=8447692183632, barcode_graphic=8447542484931 (DIFFERENT!)")

if __name__ == "__main__":
    create_correct_excel_for_layout()