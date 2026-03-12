"""Excel template generation, dummy data, and upload parsing for orders."""
import openpyxl
import tempfile
import os
import random

TMP_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.tmp')


def generate_template(variables):
    """Generate an xlsx template with column headers from variable names.
    Args:
        variables: list of dicts [{idx: int, content: str, variableName: str}, ...]
    Returns:
        filepath to the generated .xlsx in .tmp/
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for col_idx, var in enumerate(variables, start=1):
        # Use variableName if available, otherwise fall back to content
        header = var.get('variableName', '').strip() or var['content']
        ws.cell(row=1, column=col_idx, value=header)
    ws.cell(row=1, column=len(variables) + 1, value='qty')

    _write_metadata(wb, variables)

    fd, filepath = tempfile.mkstemp(suffix='.xlsx', dir=TMP_DIR)
    os.close(fd)
    wb.save(filepath)
    return filepath


def generate_dummy(variables, row_count=10):
    """Generate an xlsx with random placeholder data rows.
    Args:
        variables: list of dicts [{idx: int, content: str, variableName: str}, ...]
        row_count: number of dummy rows to generate
    Returns:
        filepath to the generated .xlsx in .tmp/
    """
    num_rows = max(1, int(row_count))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    for col_idx, var in enumerate(variables, start=1):
        # Use variableName if available, otherwise fall back to content
        header = var.get('variableName', '').strip() or var['content']
        ws.cell(row=1, column=col_idx, value=header)
    ws.cell(row=1, column=len(variables) + 1, value='qty')

    for row in range(2, num_rows + 2):
        for col_idx, var in enumerate(variables, start=1):
            ws.cell(row=row, column=col_idx, value=_random_placeholder(var['content']))
        ws.cell(row=row, column=len(variables) + 1, value=random.randint(1, 50))

    _write_metadata(wb, variables)

    fd, filepath = tempfile.mkstemp(suffix='.xlsx', dir=TMP_DIR)
    os.close(fd)
    wb.save(filepath)
    return filepath


def parse_upload(file_storage, expected_variables):
    """Parse and validate an uploaded xlsx file.
    Args:
        file_storage: werkzeug FileStorage object
        expected_variables: list of dicts [{idx: int, content: str}, ...]
    Returns:
        dict with 'success' and 'rows' or 'error'
    """
    fd, filepath = tempfile.mkstemp(suffix='.xlsx', dir=TMP_DIR)
    os.close(fd)
    file_storage.save(filepath)

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb['Data'] if 'Data' in wb.sheetnames else wb.active

        # Read metadata for index mapping
        idx_mapping = _read_metadata(wb)
        if not idx_mapping:
            # Fallback: use the idx from expected_variables
            idx_mapping = [v['idx'] for v in expected_variables]

        # Validate column count (variables + optional qty)
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        actual_cols = len([h for h in header_row if h is not None and str(h).strip() != ''])
        expected_cols_no_qty = len(expected_variables)
        expected_cols_with_qty = len(expected_variables) + 1
        has_qty = actual_cols == expected_cols_with_qty
        if actual_cols != expected_cols_with_qty and actual_cols != expected_cols_no_qty:
            return {'success': False, 'error': f'Column count mismatch: expected {expected_cols_no_qty} or {expected_cols_with_qty}, got {actual_cols}'}

        # Validate no empty headers
        for i, h in enumerate(header_row[:actual_cols]):
            if h is None or str(h).strip() == '':
                return {'success': False, 'error': f'Empty header in column {i + 1}'}

        # Read data rows and validate
        var_cols = len(expected_variables)
        rows = []
        prev_was_empty = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = list(row[:actual_cols])
            is_empty = all(v is None or str(v).strip() == '' for v in values)

            if is_empty:
                if rows:
                    prev_was_empty = True
                continue

            if prev_was_empty:
                return {'success': False, 'error': 'Empty rows between data rows are not allowed'}

            for ci, v in enumerate(values):
                if v is None or str(v).strip() == '':
                    return {'success': False, 'error': f'Empty cell at row {len(rows) + 2}, column {ci + 1}'}

            vv = {}
            for ci in range(var_cols):
                # Convert variable ID to string for consistent key format
                vv[str(idx_mapping[ci])] = str(values[ci])
            if has_qty:
                qty_val = values[var_cols]
                try:
                    qty = int(float(str(qty_val)))
                except (ValueError, TypeError):
                    return {'success': False, 'error': f'Invalid qty at row {len(rows) + 2}: {qty_val}'}
                if qty < 1:
                    return {'success': False, 'error': f'Qty must be >= 1 at row {len(rows) + 2}'}
            else:
                qty = 1
            rows.append({'variableValues': vv, 'quantity': qty})

        if not rows:
            return {'success': False, 'error': 'No data rows found'}

        return {'success': True, 'rows': rows}
    finally:
        try:
            os.unlink(filepath)
        except OSError:
            pass


def _write_metadata(wb, variables):
    """Write hidden _metadata sheet with variable index mapping."""
    meta = wb.create_sheet("_metadata")
    meta.sheet_state = 'hidden'
    meta.cell(row=1, column=1, value="col_position")
    meta.cell(row=1, column=2, value="variable_index")
    meta.cell(row=1, column=3, value="default_content")
    for row_idx, var in enumerate(variables, start=2):
        meta.cell(row=row_idx, column=1, value=row_idx - 1)
        # Store variable index as string to support both numeric and string IDs
        meta.cell(row=row_idx, column=2, value=str(var['idx']))
        meta.cell(row=row_idx, column=3, value=var['content'])


def _read_metadata(wb):
    """Read variable index mapping from _metadata sheet."""
    if '_metadata' not in wb.sheetnames:
        return None
    meta = wb['_metadata']
    mapping = []
    for row in meta.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            # Support both numeric and string variable IDs
            var_id = row[1]
            # Try to convert to int if it's a numeric string, otherwise keep as string
            try:
                var_id = int(var_id)
            except (ValueError, TypeError):
                var_id = str(var_id)
            mapping.append(var_id)
    return mapping if mapping else None


def _random_placeholder(field_name):
    """Generate contextual placeholder based on field name."""
    name_lower = field_name.lower()
    if any(w in name_lower for w in ['name', 'first', 'last']):
        return random.choice(['John', 'Jane', 'Alex', 'Maria', 'David', 'Sarah', 'Mike', 'Lisa'])
    if any(w in name_lower for w in ['address', 'street']):
        return f'{random.randint(100, 9999)} {random.choice(["Main", "Oak", "Pine", "Elm"])} St'
    if any(w in name_lower for w in ['phone', 'tel']):
        return f'555-{random.randint(100, 999)}-{random.randint(1000, 9999)}'
    if any(w in name_lower for w in ['email', 'mail']):
        return f'user{random.randint(1, 99)}@example.com'
    if any(w in name_lower for w in ['city', 'town']):
        return random.choice(['Springfield', 'Portland', 'Madison', 'Franklin'])
    if any(w in name_lower for w in ['zip', 'postal']):
        return str(random.randint(10000, 99999))
    if any(w in name_lower for w in ['title', 'position']):
        return random.choice(['Manager', 'Director', 'Engineer', 'Analyst'])
    return f'Sample {field_name} {random.randint(1, 100)}'
