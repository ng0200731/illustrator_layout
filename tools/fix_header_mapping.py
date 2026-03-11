"""
Fix the column header to variable ID mapping issue.

The problem: Your Excel has headers (barcode_digits, QR, barcode_graphic)
but the system isn't mapping them to the correct variable IDs (15, 26, 27).
"""
import json
import os

def analyze_header_mapping_issue():
    """Analyze the header mapping issue."""
    print("=== Header Mapping Issue Analysis ===")

    # Your actual Excel headers
    excel_headers = ["barcode_digits", "QR", "barcode_graphic", "qty"]

    # Expected variable mapping
    expected_mapping = {
        "barcode_digits": 15,    # Should map to variable #15
        "QR": 26,               # Should map to variable #26
        "barcode_graphic": 27   # Should map to variable #27
    }

    print("Excel Headers:")
    for i, header in enumerate(excel_headers):
        expected_var = expected_mapping.get(header, "unknown")
        print(f"  Column {i}: '{header}' → Should map to Variable #{expected_var}")

    print("\nCurrent Issue:")
    print("- barcode_digits and barcode_graphic both show same value (8447692183702)")
    print("- This suggests wrong column mapping or data duplication")

    return excel_headers, expected_mapping

def create_correct_excel_template():
    """Create Excel template with correct data structure."""
    print("\n=== Creating Correct Excel Template ===")

    # Your actual data should be:
    correct_data = {
        "barcode_digits": "8447692183702",     # Variable #15 - the main barcode
        "QR": "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116074",  # Variable #26 - QR URL
        "barcode_graphic": "8447542484929",    # Variable #27 - the header barcode (different!)
        "qty": 111
    }

    print("Correct data mapping:")
    for field, value in correct_data.items():
        if len(str(value)) > 50:
            print(f"  {field}: {str(value)[:50]}...")
        else:
            print(f"  {field}: {value}")

    # Create CSV with correct mapping
    csv_content = "barcode_digits,QR,barcode_graphic,qty\n"
    csv_content += f"{correct_data['barcode_digits']},{correct_data['QR']},{correct_data['barcode_graphic']},{correct_data['qty']}\n"

    try:
        with open('.tmp/mango_correct_mapping.csv', 'w', encoding='utf-8') as f:
            f.write(csv_content)
        print(f"\nCreated correct CSV: .tmp/mango_correct_mapping.csv")
        return True
    except Exception as e:
        print(f"Error creating CSV: {e}")
        return False

def create_layout_variable_check():
    """Create tool to check layout variable configuration."""
    print("\n=== Layout Variable Configuration Check ===")

    expected_variables = [
        {"idx": 15, "content": "barcode_digits", "variableName": "barcode_digits"},
        {"idx": 26, "content": "QR", "variableName": "QR"},
        {"idx": 27, "content": "barcode_graphic", "variableName": "barcode_graphic"}
    ]

    print("Layout should have these variables:")
    for var in expected_variables:
        print(f"  Variable #{var['idx']}: {var['variableName']} ({var['content']})")

    # Create JSON for layout check
    layout_check = {
        "variables": expected_variables,
        "note": "Ensure layout has these exact variable IDs and names"
    }

    with open('.tmp/layout_variable_check.json', 'w', encoding='utf-8') as f:
        json.dump(layout_check, f, indent=2)

    print("Created layout check: .tmp/layout_variable_check.json")

def debug_current_mapping():
    """Debug why barcode_digits and barcode_graphic show same value."""
    print("\n=== Debugging Current Mapping ===")

    print("ISSUE: Both barcode_digits and barcode_graphic show '8447692183702'")
    print("This means:")
    print("1. Either the Excel data has duplicate values in both columns")
    print("2. Or the variable mapping is pointing both variables to same column")
    print("3. Or the layout has duplicate variable references")

    print("\nTo fix:")
    print("1. Check your Excel data - ensure barcode_graphic column has different value")
    print("2. Verify variable mapping: #15→barcode_digits, #26→QR, #27→barcode_graphic")
    print("3. Check layout doesn't have duplicate variable IDs")

    print("\nExpected values:")
    print("  Variable #15 (barcode_digits): 8447692183702")
    print("  Variable #26 (QR): https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116074")
    print("  Variable #27 (barcode_graphic): 8447542484929 (should be different!)")

def create_excel_with_metadata():
    """Create Excel file with proper metadata for variable mapping."""
    print("\n=== Creating Excel with Metadata ===")

    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Headers
        ws.cell(row=1, column=1, value='barcode_digits')
        ws.cell(row=1, column=2, value='QR')
        ws.cell(row=1, column=3, value='barcode_graphic')
        ws.cell(row=1, column=4, value='qty')

        # Correct data (note: barcode_graphic should be different!)
        ws.cell(row=2, column=1, value='8447692183702')  # barcode_digits
        ws.cell(row=2, column=2, value='https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116074')  # QR
        ws.cell(row=2, column=3, value='8447542484929')  # barcode_graphic (different value!)
        ws.cell(row=2, column=4, value=111)  # qty

        # Create metadata sheet with correct variable mapping
        meta = wb.create_sheet("_metadata")
        meta.sheet_state = 'hidden'
        meta.cell(row=1, column=1, value="col_position")
        meta.cell(row=1, column=2, value="variable_index")
        meta.cell(row=1, column=3, value="default_content")

        # Correct mapping
        mappings = [
            (1, 15, "barcode_digits"),
            (2, 26, "QR"),
            (3, 27, "barcode_graphic")
        ]

        for i, (col_pos, var_idx, content) in enumerate(mappings, start=2):
            meta.cell(row=i, column=1, value=col_pos)
            meta.cell(row=i, column=2, value=var_idx)
            meta.cell(row=i, column=3, value=content)

        wb.save('.tmp/mango_fixed_excel.xlsx')
        print("Created Excel with metadata: .tmp/mango_fixed_excel.xlsx")
        print("This file has correct variable mapping and different values for each column")
        return True

    except ImportError:
        print("openpyxl not available")
        return False
    except Exception as e:
        print(f"Error creating Excel: {e}")
        return False

def main():
    """Main function to fix header mapping."""
    print("=== Column Header Mapping Fix ===")

    analyze_header_mapping_issue()
    create_correct_excel_template()
    create_layout_variable_check()
    debug_current_mapping()
    create_excel_with_metadata()

    print("\n=== SOLUTION ===")
    print("1. Use .tmp/mango_fixed_excel.xlsx - has correct data and metadata")
    print("2. Ensure barcode_graphic column has DIFFERENT value than barcode_digits")
    print("3. Check layout variables match the expected mapping")
    print("4. Variable #27 should show 8447542484929, not 8447692183702")

if __name__ == "__main__":
    main()