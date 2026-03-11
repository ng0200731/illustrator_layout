"""
Create the correct Excel file with proper metadata and different barcode values.
"""
import openpyxl

def create_correct_excel():
    """Create Excel with correct metadata and different barcode values."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Headers
    ws.cell(row=1, column=1, value='barcode_digits')
    ws.cell(row=1, column=2, value='QR')
    ws.cell(row=1, column=3, value='barcode_graphic')
    ws.cell(row=1, column=4, value='qty')

    # Sample data - barcode_graphic should be DIFFERENT from barcode_digits
    sample_data = [
        ["8447692183473", "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116073", "8447542484929", 1],
        ["8447692183702", "https://qr.mango.com/v1/p/m/2/27049218/01/01/08447692183949/21/17522116074", "8447542484930", 1]
    ]

    for i, row_data in enumerate(sample_data, start=2):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    # Create CORRECT metadata sheet
    meta = wb.create_sheet("_metadata")
    meta.sheet_state = 'hidden'
    meta.cell(row=1, column=1, value="col_position")
    meta.cell(row=1, column=2, value="variable_index")
    meta.cell(row=1, column=3, value="default_content")

    # CORRECT mapping (not 14,25,26 but 15,26,27)
    correct_mappings = [
        (1, 15, "barcode_digits"),    # Column 1 → Variable #15
        (2, 26, "QR"),               # Column 2 → Variable #26
        (3, 27, "barcode_graphic")   # Column 3 → Variable #27
    ]

    for i, (col_pos, var_idx, content) in enumerate(correct_mappings, start=2):
        meta.cell(row=i, column=1, value=col_pos)
        meta.cell(row=i, column=2, value=var_idx)
        meta.cell(row=i, column=3, value=content)

    output_file = "C:/Users/ng/Downloads/test_fixed.xlsx"
    wb.save(output_file)

    print(f"Created corrected Excel: {output_file}")
    print("Key fixes:")
    print("1. Metadata now maps to variables 15,26,27 (not 14,25,26)")
    print("2. barcode_graphic has DIFFERENT values from barcode_digits")
    print("3. Variable #15: barcode_digits (8447692183473)")
    print("4. Variable #26: QR URL (...6073)")
    print("5. Variable #27: barcode_graphic (8447542484929) ← DIFFERENT!")

if __name__ == "__main__":
    create_correct_excel()